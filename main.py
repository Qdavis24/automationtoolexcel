import pandas
from openpyxl import load_workbook
from docx import Document


def remove_na_questions(drop_index, series):
    # identify values with green bg
    series = series.drop(drop_index)
    return series[series.str[:3] != "see"].dropna().tolist()


try:
    dataframe = pandas.read_excel("./files/input.xlsx", header=2)
    wb = load_workbook("./files/input.xlsx")
    ws = wb.active
    ws_col = ws["k"]
    df = dataframe.iloc[:, [2, 10]]
    df.columns = ["section", "question"]
    drop_index = [cell.row - 1 for cell in ws_col if cell.fill.start_color.rgb == "FFC6EFCE"]

    diff_sections = df["section"]
    diff_sections = diff_sections.dropna()
    diff_sections = diff_sections.unique()

    output = {}

    for section in diff_sections:
        strn = section[:4]
        output[strn] = remove_na_questions(drop_index, df[df["section"] == section]["question"])
    print(output)
except FileNotFoundError as error:
    print(f"{error}: excel file path specified is not correct")

else:
    try:
        looking = True
        curr_section = ""
        document = Document("./files/template.docx")
        j = 0
        for par in document.paragraphs:
            original_lines = par.text.split("\n")
            target = 0

            for i in range(0, len(original_lines)):
                if looking:
                    if original_lines[i] == "Confirm/Submit/Describe…":
                        looking = False
                        replace_string = ""
                        for k, el in enumerate(output[curr_section]):
                            j += 1
                            if k != 0:
                                replace_string += "•"
                            replace_string += f"{j}. {el}\n\n[enter response here]\n\n"
                        original_lines[i] = replace_string
                        break
                if original_lines[i][9:13] in output:
                    curr_section = original_lines[i][9:13]
                    looking = True
            output_string = "".join(original_lines)
            par.text = output_string

    except FileNotFoundError as error:
        print(f"{error}: Template file path specified is incorrect")
    else:
        document.save("./files/output.docx")
